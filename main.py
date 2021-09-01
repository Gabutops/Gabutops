from datetime import datetime
import random
from openpyxl import load_workbook
import logging
import typing
import PyQt5
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QImage, QRegExpValidator, QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient,QMovie
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog, QMainWindow, QInputDialog, QHeaderView, QMessageBox, QPushButton, QWidget,  QLineEdit, QTableWidget, QTableWidgetItem, QItemDelegate, QLabel
from PyQt5.uic import loadUi
from PyQt5.QtCore import QCoreApplication, QPropertyAnimation, QDateTime, QMetaObject, QObject, QPoint, QRect, QSize, QTime, QUrl, Qt, QEvent,QTimer, pyqtSignal, pyqtSlot
import sys
import os
import json
import time
import selenium
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome import options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import MoveTargetOutOfBoundsException, TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select

USER = ""
SENHA = ""
CONT = 1
MAXCONT = 1
EXCEL = ""


class Worker(QObject):
    global EXCEL
    global MAXCONT
    global CONT
    loginSender = pyqtSignal(bool)
    outputSender = pyqtSignal(str)
    def __init__(self, driver, parent = None):
        super(Worker, self).__init__(parent)
        self.driver = driver
        log = "{}.log".format(datetime.now().strftime("%d-%m-%Y"))
        with open(r'logs\{}'.format(log),'a') as f:
            logging.basicConfig(filename=log, level=logging.INFO)
    def start(self):
        self.driver.get('http://wss.credisim.com.br/BSGWEBSITES/WebAutorizador/Login/AC.UI.LOGIN.aspx')
        url = self.driver.current_url.split('?')
        for session in range(len(url)):
            if str(url[session]).startswith('FISession'):
                self.session = url[session]
                print(self.session)
                logging.info("session: {}".format(self.session))
            else:
                print(url[session])
        self.login(True)
    def loadExcel(self,excelUrl):
        if excelUrl == "":
            pass
        else:
            wb2 = load_workbook(excelUrl)
            self.excel = wb2.active
            logging.info("excel: {}({})".format(self.excel.title,excelUrl))
            MAXCONT = self.excel.max_row
            logging.info('MAXCONT={}'.format(self.excel.max_row))

    @pyqtSlot(bool)
    def login(self,lg):
        self.popupLogin(lg)
    def popupLogin(self,lg):
        if lg == True:
            self.loginSender.emit(True)
            logging.info("LoginScreen: {}".format('Open'))
        else:
            self.loginSender.emit(False)
            logging.info("LoginScreen: {}".format('Closed'))
    def input_login(self, loginText):
        inputLogin = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table[2]/tbody/tr/td[10]/table/tbody/tr/td[1]/input')
            ))
        while not inputLogin.get_attribute('value') == str(loginText):
            if int(len(inputLogin.get_attribute('value'))) >= int(len(str(loginText))):
                  for caracteres in inputLogin.get_attribute('value'):
                        if caracteres == 0:
                            pass
                        else:
                            inputLogin.send_keys(Keys.BACK_SPACE)
            elif inputLogin.get_attribute('value') != loginText :
                inputLogin.send_keys(loginText)
            if inputLogin.get_attribute('value') == loginText :
                print("{}(userOk)".format(inputLogin.get_attribute('value')))
                logging.info("{}(userOk)".format(inputLogin.get_attribute('value')))
    def input_senha(self, senhaText):
        senha = True
        while senha:
            try:
                inputSenha =  WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table[2]/tbody/tr/td[14]/table/tbody/tr/td[1]/input')
                    ))
                if inputSenha.get_attribute('value') == "":
                    inputSenha.send_keys(str(senhaText))
                elif inputSenha.get_attribute('value') != str(senhaText) :
                    for caracteres in range(len(inputSenha.get_attribute('value'))):
                        if inputSenha.get_attribute('value') == "":
                            break
                        else:
                            inputSenha.send_keys(Keys.BACK_SPACE)
                            print('senhaClean')

                elif inputSenha.get_attribute('value') == senhaText:
                    print("{}(senhaOk)".format(str(inputSenha.get_attribute('value'))))
                    logging.info("{}(senhaOk)".format(str(inputSenha.get_attribute('value'))))
                    senha = False
                else:
                    print('senhaError:'
                    '\nsenhaColocada: {}'
                    '\nsenhaPosta: {}'.format(str(senhaText,inputSenha.get_attribute('value'))))
                    logging.error('senhaError: {}'.format(str(senhaText,inputSenha.get_attribute('value'))))
            except StaleElementReferenceException as e:
                print(e.args,'(senhaException)')
                logging.error("senhaException;{}".format(e.args))
    def entrarClick(self,click):
        while click:
            try:
                btn = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table[2]/tbody/tr/td[16]/a')
                  ))
                btn.click()
                try:
                    alert = WebDriverWait(self.driver, 5).until(EC.alert_is_present())
                    text = alert.text
                    if text != "Senha Incorreta.":
                        self.errorSenha(text)
                        self.cleanSenha(True)
                        alert.accept()
                        click = False
                    else:
                        print(text)
                        alert.accept()
                        click = False
                except TimeoutException as e:
                    print(e.args,'(noAlert)')
                    site = 'http://wss.credisim.com.br/BSGWEBSITES/WebAutorizador/'
                    if self.driver.current_url == "http://wss.credisim.com.br/BSGWEBSITES/WebAutorizador/?{}".format(self.session):
                        click = False
                        print('click:{}(LoginOk)'.format(click))
                        logging.info("login: ok")
                        self.popupLogin(False)
                    else:
                        click = True
            except Exception as e:
                print(e.args,'(clickException)')
                logging.error("clickException;{}".format(e.args))
                click = False
    def cleanSenha(self,clean):
        print('this will clean')
    def errorSenha(self,error):
        print(error)
    def gotoTermo(self,):
        url = 'http://wss.credisim.com.br/BSGWEBSITES/WebAutorizador/MenuWeb/INSS/DadosBeneficiario/UI.CD.DadosBeneficiario.aspx?{}'.format(self.session)
        if not self.driver.currenturl == url:
            self.driver.get(url)
        else:
            print('sucess(getOk)')
    def impressaoTermo(self,lap):
        while lap:
            self.cpfTermo(self.excel.cell(CONT,2).value)
            self.nomeTermo(self.excel.cell(CONT,3).value)
    def cpfTermo(self,cpf):
        cpfInput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[1]/tbody/tr[3]/td/table/tbody/tr/td[1]/input')
            ))
        while not cpfInput.get_attribute('value') == cpf:
            if cpfInput.get_attribute('value') == "":
                cpfInput.send_keys(cpf)
            elif cpfInput.get_attribute('value') != cpf:
                for caracters in range(len(cpfInput.get_attribute('value'))):
                    if caracters == 0:
                        pass 
                    cpfInput.send_keys(Keys.BACK_SPACE)
                print('cpfInput_cleaned')
            elif cpfInput.get_attribute('value') == cpf:
                print('{}(cpfOk)'.format(cpfInput.get_attribute('value')))
                 
    def nomeTermo(self,nome):
        nomeInput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[1]/tbody/tr[3]/td/table/tbody/tr/td[2]/input')
            ))
        while not nomeInput.get_attribute('value') == nome:
            self.nome = nome
            if nomeInput.get_attribute('value') == "":
                nomeInput.send_keys(nome)
            elif nomeInput.get_attribute('value') != nome:
                for caracters in range(len(nomeInput.get_attribute('value'))):
                    if caracters == 0:
                        pass 
                    nomeInput.send_keys(Keys.BACK_SPACE)
                print('nomeInput_cleaned')
            elif nomeInput.get_attribute('value') == nome:
                print('{}(nomeOk)'.format(nomeInput.get_attribute('value')))
                self.output('{}(nome)'.format(nomeInput).get_attribute('value'))
    def telefone(self):
        SIM = "11942032151"
        DDD = SIM[:2]
        TEL = SIM[9:]
        dddInput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[1]/tbody/tr[3]/td/table/tbody/tr/td[3]/input[1]')
            ))
        while not dddInput.get_attribute('value') == DDD:
            if dddInput.get_attribute('value') == "":
                dddInput.send_keys(DDD)
            elif dddInput.get_attribute('value') != TEL:
                for caracters in range(len(dddInput.get_attribute('value'))):
                    if caracters == 0:
                        pass
                    else:
                        print(dddInput.send_keys(Keys.BACK_SPACE))
            else:
                print(dddInput.get_attribute('value'))
        telInput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[1]/tbody/tr[3]/td/table/tbody/tr/td[3]/input[2]')
            ))
        while not telInput.get_attribute() == TEL:
            if telInput.get_attribute('value') == "":
                telInput.send_keys(TEL)
            elif telInput.get_attribute('value') != TEL:
                for caracters in range(len(telInput.get_attribute('value'))):
                    telInput.send_keys(Keys.BACK_SPACE)
            else:
                print(telInput.get_attribute('value'))
    def emailTermo(self):
        email = (str(self.nome).replace(" ","_")+"@gmail.com")
        emailInput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[1]/tbody/tr[3]/td/table/tbody/tr/td[4]/input')
            ))
        while not emailInput.get_attribute('value') == email:
            if emailInput.get_attribute('value') =="":
                emailInput.get_attribute('value').send_keys(Keys.BACK_SPACE)
            elif emailInput.get_attribute('value') != email:
                for c in range(len(emailInput.get_attribute('value'))):
                    emailInput.send_keys(Keys.BACK_SPACE)
            else:
                print(emailInput.get_attribute('value'))
    def localTermo(self,local):
        localInput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[1]/tbody/tr[4]/td/table/tbody/tr/td[1]/input')
            ))
        while not localInput.get_attribute('value') == local:
            if localInput.get_attribute('value') == "":
                localInput.send_keys(local)
            elif localInput.get_attribute('value') != local:
                for i in range(len(localInput.get_attribute('value'))):
                    localInput.send_keys(Keys.BACK_SPACE)
            else:
                print(localInput.get_attribute('value'))
    def submitTermo(self,click):
        while click:
            try:
                btn = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td/div/div/div[2]/div[1]/div/table[2]/tbody/tr/td[2]/span/div[1]/div[2]/div[2]/table/tbody/tr/td')
                    ))
                btn.click()
                try:
                    alert = WebDriverWait(self.driver,20).until(EC.alert_is_present())
                    text = alert.text
                    if alert == 'Termo de Autorização enviado para Assinatura Digital com sucesso!':
                        alert.accept()
                        print(text)
                        click = False
                    else:
                        print(text)
                        alert.accept()
                        click = False
                except TimeoutException:
                    print('no alert')
            except Exception as e:
                print(e,e.args)
    @pyqtSlot(str)
    def output(self,text):
        self.output(text)
    def configOutput(self,text):
        self.outputSender.emit(text)


class Assinatura(QObject):
    def __init__(self,driver, parent = None):
        super(Assinatura).__init__(parent)
        self.driver =driver
    def receiveSMS(self):
        'aqui ira vir a função de receber o SMS'
        'e encaminhar para a função SMS'
    def SMS(self):
        smsText = ""
        link = smsText.split('link de consulta:')[-1]
        self.getLink(link)
    def getLink(self, link):
        self.driver.get(link) 
    def assinaCerto(self):
        pagina_Termo = self.driver.current_window_handle
        acao = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/form/div[6]/div/div/div[1]/div[2]/div/div/div/table/tbody/tr/td[1]/a[1]')
                    ))
        acao.click()
        try:
            chwd = self.driver.window_handles
            for w in chwd:
            #switch focus to child window
                if(w!=pagina_Termo):
                    self.driver.switch_to.window(w)
        except Exception as e:
            print(e)
    def receiveCodigo(self):
        "aqui ira receber o codigo do SMS (o segundo SMS)"
    def inputCodigo(self,codigo):
        accept = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[3]/div/button')
                    ))
        accept.click()

        input = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/div[2]/input')
                    ))
        while not input.get_attribute('value') == codigo:
            if input.get_attribute('value') =="":
                input.send_keys(codigo)
                confirmar = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/button[1]')
                    ))
                confirmar.click()
            elif input.get_attribute('value') != codigo:
                for caracters in range(len(input.get_attribute('value'))):
                    input.send_keys(Keys.BACK_SPACE)
            else:
                print(input.get_attribute('value'))
    def assinar(self):
        antW = 0
        d = True
        while d:
            try:
                canvas = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[2]/canvas')
                            ))
                drawing = ActionChains(self.driver)\
                        .click_and_hold(canvas)
                for i in range(random.randint(30,35)):
                    randW = random.uniform(0, 70)
                    randH = random.uniform(0, 50)
                    randP = random.uniform(1,-1)
                    
                    w = (randW*randP)
                    if w+antW >= 75:
                        w = antW *-(random.uniform(0.2, 1))
                        print('corrigindo W: {}'.format(w))
                    elif w - antW<= -75:
                        w = antW *-(random.uniform(0.2, 1))
                        print('corrigindo W: {}'.format(w))
                    h = (randH*randP)
                    
                    drawing.move_by_offset(h, w)
                    antW = (w)
                    
                
                drawing.release()
                drawing.perform()
                print('ass.: robo :)')
                d = False
            except MoveTargetOutOfBoundsException as e:
                print(e.msg)
                print('fora da caixa')  
    def confirmar(self):
        confirmarBtn =  WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/button[1]')
                            ))
        if confirmarBtn.get_attribute('disabled') == 'false':
            print('assinatura necessaria')
            self.assinar()
        else:
            confirmarBtn.click()
    def alerta(self):
        alertaBox = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]')
                            ))
        alerta = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div[2]')
                            ))
        print(alerta.get_attribute('textContent'))
        while alerta.is_enabled():
            sim = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div[3]/div/button[1]')
                            ))
            sim.click()
            alertaBox2 = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]')
                            ))
            while alertaBox2.is_enabled():
                ok = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div[3]/div/button')
                            ))
                alerta2 = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div[2]')
                            ))
                print(alerta2.get_attribute('textContent'))
                ok.click()
                time.sleep(1)
class MainWindow(QMainWindow):
    global EXCEL
    def __init__(self) -> None:
        super().__init__()
        def closeEvent(self, event):
            print('Calling')
            print('event: {0}'.format(event))
            event.accept()
        loadUi('templates\inteface.ui',self)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("images\icon.png"), QtGui.QIcon.Selected, QtGui.QIcon.On)
        self.setWindowIcon(icon)
        self.setWindowIconText('Portablidade Banco Seguro')
        self.setWindowTitle('Portablidade Banco Seguro')
        
        thread = QtCore.QThread(self)
        thread.start()


        download_dir=(r'\\lpc-svr-02\Vendas\JSON\JSON-BANCOSEGURO')
        options = Options()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_experimental_option("prefs", {
              "download.default_directory": download_dir,
              "download.prompt_for_download": False,
              "download.directory_upgrade": True,
              "safebrowsing.enabled": False,
              "plugins.always_open_pdf_externally" : True
            })
            #options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-device-discovery-notifications')
        options.add_argument("--disable-extensions")
        self.worker = Worker(driver = webdriver.Chrome(options = options))
        
        self.worker.moveToThread(thread)

        #CONNECTS*

        self.comecar.clicked.connect(self.worker.start)
        self.worker.loginSender.connect(self.Popup)
        self.pushButton.clicked.connect(self.browsefiles)
    def Popup(self,appear):
        global USER
        global SENHA
        if not self.lineEdit.text() == "":
            if appear is True:
                self.popup = loadUi(r"templates\login.ui")
                self.popup.setModal(True)
                self.popup.lineEdit_login.editingFinished.connect(self.enableButton)
                self.popup.lineEdit_login.editingFinished.connect(self.login)
                self.popup.lineEdit_senha.textChanged.connect(self.enableButton)
                self.popup.lineEdit_senha.editingFinished.connect(self.senha)
                self.popup.acessar.clicked.connect(self.entrar)
                self.popup.exec()
            else:
                self.popup.close()
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Information)
                msgBox.setText("Login ok")
                msgBox.setWindowTitle("Login")
                msgBox.setStandardButtons(QMessageBox.Ok)
                returnValue = msgBox.exec()
                if returnValue == QMessageBox.Ok:
                    print('OK clicked')
                    self.start(True)
        else:
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Information)
            msgBox.setText("Erro: Selecione uma pasta")
            msgBox.setWindowTitle("Path not found")
            msgBox.setStandardButtons(QMessageBox.Ok)
            returnValue = msgBox.exec()
            if returnValue == QMessageBox.Ok:
                print('OK clicked')
    def login(self):
        USER = self.popup.lineEdit_login.text()
        self.input_login(USER)
    def senha(self):
        SENHA = self.popup.lineEdit_senha.text()
        self.input_senha(SENHA)
    def input_login(self,user):
        self.worker.input_login(user)
        print(user)
    def input_senha(self,psw):
        self.worker.input_senha(psw)
        print(psw)
    def entrar(self):
        self.worker.entrarClick(True)
    def enableButton(self):
        if len(self.popup.lineEdit_login.text()) > 0 and len(self.popup.lineEdit_senha.text()) > 0:
                self.popup.acessar.setEnabled(True)
    def browsefiles(self):
        global url
        raiz =  os.path.join(os.environ["HOMEPATH"], "Desktop")
        fname = QFileDialog.getOpenFileName(self,'Open File',raiz,'*xlsx')
        url = fname[0]
        self.lineEdit.setText(url)
        EXCEL = url
        if not url =="":
            self.worker.loadExcel(EXCEL)
        return url
    def start(self,go):
        if go is True and CONT >=MAXCONT:
            self.worker.impressaoTermo(True)
        else:
            self.worker.impressaoTermo(False)
            

    
        









app = QApplication(sys.argv)
mainwindow = MainWindow()
widget = QtWidgets.QStackedWidget()
widget.addWidget(mainwindow)
widget.show()
sys.exit(app.exec_())